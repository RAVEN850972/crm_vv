import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from django.http import HttpResponse
from customer_clients.models import Client
from orders.models import Order, OrderItem
from services.models import Service
from finance.models import Transaction

def export_clients_to_excel():
    """Экспорт клиентов в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Заголовки
    headers = ['ID', 'Имя', 'Адрес', 'Телефон', 'Источник', 'Дата создания']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Данные
    clients = Client.objects.all().order_by('id')
    for row_num, client in enumerate(clients, 2):
        source_display = dict(Client.SOURCE_CHOICES).get(client.source, client.source)
        
        ws.cell(row=row_num, column=1, value=client.id)
        ws.cell(row=row_num, column=2, value=client.name)
        ws.cell(row=row_num, column=3, value=client.address)
        ws.cell(row=row_num, column=4, value=client.phone)
        ws.cell(row=row_num, column=5, value=source_display)
        ws.cell(row=row_num, column=6, value=client.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'clients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

def export_orders_to_excel():
    """Экспорт заказов в Excel"""
    wb = openpyxl.Workbook()
    
    # Лист с заказами
    ws_orders = wb.active
    ws_orders.title = "Заказы"
    
    # Заголовки для заказов
    order_headers = ['ID', 'Клиент', 'Менеджер', 'Статус', 'Общая стоимость', 'Дата создания', 'Дата завершения']
    for col_num, header in enumerate(order_headers, 1):
        ws_orders.cell(row=1, column=col_num, value=header)
    
    # Данные заказов
    orders = Order.objects.all().order_by('id')
    for row_num, order in enumerate(orders, 2):
        status_display = dict(Order.STATUS_CHOICES).get(order.status, order.status)
        completed_at = order.completed_at.strftime('%Y-%m-%d %H:%M:%S') if order.completed_at else ''
        
        ws_orders.cell(row=row_num, column=1, value=order.id)
        ws_orders.cell(row=row_num, column=2, value=order.client.name)
        ws_orders.cell(row=row_num, column=3, value=order.manager.get_full_name())
        ws_orders.cell(row=row_num, column=4, value=status_display)
        ws_orders.cell(row=row_num, column=5, value=float(order.total_cost))
        ws_orders.cell(row=row_num, column=6, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws_orders.cell(row=row_num, column=7, value=completed_at)
    
    # Лист с позициями заказов
    ws_items = wb.create_sheet("Позиции заказов")
    
    # Заголовки для позиций
    item_headers = ['ID заказа', 'Услуга', 'Категория', 'Цена', 'Продавец', 'Дата создания']
    for col_num, header in enumerate(item_headers, 1):
        ws_items.cell(row=1, column=col_num, value=header)
    
    # Данные позиций
    items = OrderItem.objects.all().order_by('order__id')
    for row_num, item in enumerate(items, 2):
        category_display = dict(Service.CATEGORY_CHOICES).get(item.service.category, item.service.category)
        
        ws_items.cell(row=row_num, column=1, value=item.order.id)
        ws_items.cell(row=row_num, column=2, value=item.service.name)
        ws_items.cell(row=row_num, column=3, value=category_display)
        ws_items.cell(row=row_num, column=4, value=float(item.price))
        ws_items.cell(row=row_num, column=5, value=item.seller.get_full_name())
        ws_items.cell(row=row_num, column=6, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Автоподбор ширины колонок для всех листов
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

def export_finance_to_excel():
    """Экспорт финансовых операций в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Финансы"
    
    # Заголовки
    headers = ['ID', 'Тип', 'Сумма', 'Описание', 'Связанный заказ', 'Дата создания']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Данные
    transactions = Transaction.objects.all().order_by('-created_at')
    for row_num, transaction in enumerate(transactions, 2):
        type_display = dict(Transaction.TYPE_CHOICES).get(transaction.type, transaction.type)
        order_id = transaction.order.id if transaction.order else ''
        
        ws.cell(row=row_num, column=1, value=transaction.id)
        ws.cell(row=row_num, column=2, value=type_display)
        ws.cell(row=row_num, column=3, value=float(transaction.amount))
        ws.cell(row=row_num, column=4, value=transaction.description)
        ws.cell(row=row_num, column=5, value=order_id)
        ws.cell(row=row_num, column=6, value=transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'finance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response